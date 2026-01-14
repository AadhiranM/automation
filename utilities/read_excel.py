import datetime

import openpyxl


def get_test_data(file_path, sheet_name):
    import openpyxl
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    # Read header
    headers = [cell.value.strip() if isinstance(cell.value, str) else cell.value for cell in sheet[1]]

    test_data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if all(cell is None for cell in row):
            continue

        row_data = {}
        for header, cell in zip(headers, row):
            if isinstance(cell, str):
                row_data[header] = cell.strip()
            elif isinstance(cell, (datetime.datetime, datetime.date)):
                row_data[header] = cell.strftime("%d-%m-%Y")  # convert to string
            else:
                row_data[header] = "" if cell is None else str(cell)
        test_data.append(row_data)

    return test_data

#
# def get_test_data(file_path, sheet_name):
#     data = []
#     wb = openpyxl.load_workbook(file_path)
#     sheet = wb[sheet_name]
#
#     headers = [cell.value for cell in sheet[1]]
#
#     for row in sheet.iter_rows(min_row=2, values_only=True):
#         if not any(row):   # 🔥 THIS LINE FIXES data2
#             continue
#
#         row_dict = dict(zip(headers, row))
#         data.append(row_dict)
#
#     return data
