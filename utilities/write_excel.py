from openpyxl import load_workbook

def write_test_result(file_path, sheet_name, row, result):
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]

    # Assuming Result column is 3
    sheet.cell(row=row, column=3).value = result

    workbook.save(file_path)
